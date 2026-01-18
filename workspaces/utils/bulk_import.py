from __future__ import annotations
import csv
import re
from io import StringIO, BytesIO
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook, Workbook


EXCLUDE_TYPES = {"SHAPE", "STATIC_TEXT", "BARCODE", "QRCODE", "SERIAL"}  # not user-input
REQUIRED_COL = "EAN_CODE"
OPTIONAL_COL = "GS1_CODE"
QTY_COL = "QUANTITY"
MIN_QTY = 1
MAX_QTY = 100


def norm_header(h: str) -> str:
    h = (h or "").strip()
    h = re.sub(r"\s+", "_", h)
    h = re.sub(r"[^A-Za-z0-9_]+", "", h)
    return h.upper()


def build_expected_headers(layout_items: List[dict]) -> Tuple[List[str], List[str]]:
    """
    Returns:
      headers: [EAN_CODE, GS1_CODE, <var1>, <var2>...]
      var_keys: [<var1>, <var2>...]
    Uses template keys as-is for var keys.
    """
    var_keys = []
    seen = set()

    for it in (layout_items or []):
        ft = (it.get("field_type") or "").upper()
        key = (it.get("key") or "").strip()
        if not key:
            continue
        if ft in EXCLUDE_TYPES:
            continue
        # This is a user-input field
        if key not in seen:
            var_keys.append(key)
            seen.add(key)

    headers = [REQUIRED_COL, OPTIONAL_COL, QTY_COL] + var_keys
    return headers, var_keys


def parse_csv_bytes(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    f = StringIO(text)
    reader = csv.DictReader(f)
    headers = reader.fieldnames or []
    rows = []
    for r in reader:
        if not any((v or "").strip() for v in (r or {}).values()):
            continue
        rows.append({k: (v or "").strip() for k, v in (r or {}).items()})
    return headers, rows


def parse_xlsx_bytes(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    raw_headers = [str(x or "").strip() for x in all_rows[0]]
    headers = raw_headers
    rows = []
    for row in all_rows[1:]:
        if not row:
            continue
        values = [str(x or "").strip() for x in row]
        if not any(values):
            continue
        d = {}
        for i, h in enumerate(raw_headers):
            if not h:
                continue
            d[h] = values[i] if i < len(values) else ""
        rows.append(d)
    return headers, rows


def make_csv_template_bytes(headers: List[str]) -> bytes:
    out = StringIO()
    w = csv.writer(out)
    w.writerow(headers)
    return out.getvalue().encode("utf-8")


def make_xlsx_template_bytes(headers: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def validate_and_normalize_rows(
    expected_headers: List[str],
    var_keys: List[str],
    file_headers: List[str],
    rows: List[Dict[str, str]],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Returns (normalized_rows, errors)

    normalized_rows is:
      [
        {"ean_code": "...", "gs1_code": "...", "field_values": {...}},
        ...
      ]
    """

    errors = []

    if not rows:
        return [], ["Uploaded file has no data rows. Please fill at least 1 row."]

    # Build mapping: normalized header -> original header
    file_map = {}
    for h in (file_headers or []):
        nh = norm_header(h)
        if nh and nh not in file_map:
            file_map[nh] = h

    expected_norm = [norm_header(h) for h in expected_headers]

    # Required: EAN_CODE present
    if REQUIRED_COL not in file_map:
        errors.append("Missing EAN_CODE column. Please download the correct template and re-upload.")

    if QTY_COL not in file_map:
        errors.append("Missing QUANTITY column. Please download the correct template and re-upload.")

    # Optional: GS1_CODE (ok if missing)
    # All variable keys must exist (case-insensitive)
    for k in var_keys:
        nk = norm_header(k)
        if nk not in file_map:
            errors.append(f"Missing column: {k}. Please download the correct template and re-upload.")

    if errors:
        return [], errors

    # Normalize rows
    normalized = []
    for idx, r in enumerate(rows, start=1):
        ean = (r.get(file_map.get(REQUIRED_COL, ""), "") or "").strip()
        gs1 = (r.get(file_map.get(OPTIONAL_COL, ""), "") or "").strip() if file_map.get(OPTIONAL_COL) else ""
        qty_raw = ""

        if file_map.get(QTY_COL):
            qty_raw = (r.get(file_map.get(QTY_COL), "") or "").strip()

        # default blank to 1 (safe)
        if not qty_raw:
            qty = 1
        else:
            try:
                qty = int(qty_raw)
            except ValueError:
                errors.append(f"Row {idx}: QUANTITY must be a whole number (1–{MAX_QTY}).")
                continue

        if qty < MIN_QTY or qty > MAX_QTY:
            errors.append(f"Row {idx}: QUANTITY must be between {MIN_QTY} and {MAX_QTY}.")
            continue


        if not ean:
            errors.append(f"Row {idx}: EAN_CODE is empty. Barcode/QR cannot be generated.")
            continue

        fv = {}
        for k in var_keys:
            orig_header = file_map.get(norm_header(k))
            fv[k] = (r.get(orig_header, "") or "").strip() if orig_header else ""

        normalized.append({
            "ean_code": ean,
            "gs1_code": gs1,
            "quantity": qty,
            "field_values": fv,
        })

    if errors:
        return [], errors

    return normalized, []
